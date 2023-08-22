#!/usr/bin/env python3
import openpyxl
import sys
import math
from openpyxl import Workbook

def create_new_files(path, header_lines, file_lines):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.worksheets[0]

    num_files = math.ceil((sheet.max_row - int(header_lines)) / int(file_lines))

    for i in range(1, num_files + 1):
        if i >= 10:
            name = path[:-5] + str(i) + ".xlsx"
        else:
            name = path[:-5] + "0" + str(i) + ".xlsx"

        new_workbook = Workbook()
        new_sheet = new_workbook.active

        for j in range(1, int(header_lines) + 1):
            for k in range(1, sheet.max_column + 1):
                data = sheet.cell(row = j, column = k)
                new_sheet.cell(row = j, column = k).value = data.value

        for x in range(int(header_lines) + 1, 
                       (int(header_lines) + 1) + int(file_lines)):
            for y in range(1, sheet.max_column + 1):
                data = sheet.cell(row = x + (int(file_lines) * (i-1)), column = y)
                new_sheet.cell(row = x, column = y).value = data.value
            

        new_workbook.save(name)


create_new_files(sys.argv[1], sys.argv[2], sys.argv[3])

